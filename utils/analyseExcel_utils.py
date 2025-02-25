"""
文件名: analyseExcel_utils.py
作者: LvMC
创建时间: 2024-10-12
版本: 1.0

文件说明： 解析excel文本用例 为指定json格式的数据。
json数据结构： 最外层列表包含 > 多个模块list > 每个模块包含多个接口list > 每个接口包含多个用例list。
获取到的数据结构如下：
[
	{
	模块1各字段键: 对应值,
	"interface_lists": [
	  {
		接口1各字段键: 对应值,
		接口1"case_lists":[ {接口1-用例1数据},{接口1-用例2数据} ]
	  },
	 {
		接口2各字段键: 对应值,
		接口2"case_lists":[ {接口2-用例1数据},{接口2-用例2数据} ]
	  }
    ]
},
    { 模块2对应数据}
]
"""


import openpyxl
from utils.logger import Logger


class AnalyseExcel:


    def __init__(self,excel_file):
        # 加载Excel表格，获取'controller'表中的有效列数
        try:
            self.wb = openpyxl.load_workbook(excel_file) #加载 Excel 文件
            self.controller = self.wb['controller'] #读取名为'controller'的工作表
        except Exception:
            Logger.error("【异常提醒】加载{}文件错误".format(excel_file))
        # self.controller = self.wb['controller']
        for cell in self.controller[1]: #遍历'controller'表的第一行，并获取该行中的有效(非空)列数
            if cell.value == None:
                self.length_of_controller = cell.column - 1
                break

    def read_module_data(self):
        # 从 controller 工作表中读取模块数据
        # 数据结构：[{模块1各字段键:对应值, "interface_lists": [{接口1数据}{接口2数据}]},  {模块2各字段键:对应值,"interface_lists": [{接口1数据}{接口2数据}]}]
        module_list = [] # [{模块1数据},{模块2数据}]
        for row in self.controller.iter_rows(min_row=2,min_col=1,max_col=self.length_of_controller): #从第二行第一列 开始遍历所有行
            # row 是一个元组，元组下标是从0开始。
            if row[0].value is not None: #判断第一列是否有值：有则为模块数据行
                module_dict = {} #上一行代码已筛选出一行行的模块数据行，故此处所处层级是对当行数据的操作。每一行的所有单元格要聚合 存储到一个字典中，故在此处定义一个字典
                for cell in row: #遍历时，因为要通过相对位置的单元格中是否有值来筛选数据，故可先获取：当前cell值、相对位置celL值。
                    flag = cell.value
                    next_flag = self.controller.cell(cell.row+1,cell.column).value #获取同列下一行的数据 值
                    if flag is not None or next_flag is None: #符合此条件的为模块数据；否则为接口数据
                        module_dict.update({self.controller.cell(1,cell.column).value:cell.value}) #{键：值}={与当前单元格同列第一行的单元格值：当前单元格值}
                    else:
                        # 此处所处层级是对当行数据的操作。上面的for循环已经遍历把 模块相关的其它字段添加到字典中，该行中的其它字段为接口相关字段，用 module_dict['interface_list']存放接口字段。
                        module_dict['interface_list'] = self.read_interface_data(row[3].value,cell.row+1,cell.column,self.length_of_controller) #因为后续读取用例时，需用到module_code判断读哪个用例表，故需传入module_code。
                        break
                module_list.append(module_dict) # (24/10/12)：此处所处层级是对当行数据的操作。前面已经把一行中所有单元格都聚合到字典了。可在此处把聚合完成的字典append到列表中。
            else:
                Logger.error("【异常提醒】单元格[{0},{1}]值为空，请检查controller表数据".format(row[0].row,row[0].column))
        return module_list


    def read_interface_data(self,module_code,min_row,min_col,max_col):
        # 从 controller 工作表中读取接口数据
        # 数据结构："interface_lists": [ {接口1各字段键:对应值, 'case_lists':[ {接口1-用例1数据},{接口1-用例2数据} ] },  {接口2各字段键:对应值, 'case_lists':[{接口2-用例1数据}{接口2-用例2数据}]} ]
        interface_list = [] #[{模块1各字段键:对应值, "interface_lists": [{接口1数据}{接口2数据}]},  { 模块2各字段键:对应值,"interface_lists": [{接口1数据}{接口2数据}] }]
        for row in self.controller.iter_rows(min_row=min_row,min_col=min_col,max_col=max_col): # 遍历接口数据各行。min_row、min_col在读取模块数据时可推测
            if row[0].value is None:  #读取到当前行的第一个单元格值为空时，则该行已不是接口数据行
                break
            else: # 此处筛选出的都为接口数据行
                interface_dict = {} #(24/10/12)：此处所处层级是对当行数据的操作；一个字典存放一个接口数据。
                for cell in row:
                    interface_dict.update({self.controller.cell(1,cell.column).value:cell.value})
                interface_dict['case_lists'] = self.read_case_data(module_code,interface_code=row[1].value) #(24/10/12)：上面for循环已经把一行中其它字段添加到字典中了，故该行中的case_lists字段在此处添加一次。
                interface_list.append(interface_dict) # (24/10/12)：此处所处层级是对当行数据的操作。前面已经把一行中所有单元格都聚合到字典了。可在此处把聚合完成的字典append到列表中。
        return interface_list


    def read_case_data(self,module_code,interface_code):
        # 从 controller 工作表中读取用例数据
        # 数据结构：
        case_lists = []
        case_sheet = self.wb[str(module_code)]
        for cell in case_sheet[1]:
            if cell.value == None:
                length_of_case_sheet = cell.column - 1

        for item in case_sheet['A']:
            if item.value == interface_code:
                for row in case_sheet.iter_rows(min_row=item.row+1,min_col=item.column+1,max_col=length_of_case_sheet):
                    case_dict = {} #(24/10/12)：此处所处层级是对当行数据的操作；且此条件中处理的每一行数据分别要聚合到一个字典中。故此处需定义字典
                    if row[0].value is None:
                        break
                    else:
                        for cell in row:
                            case_dict.update({case_sheet.cell(1,cell.column).value:cell.value})
                        case_lists.append(case_dict) # (24/10/12)：此处所处层级是对当行数据的操作。前面已经把一行中所有单元格都聚合到字典了。可在此处把聚合完成的字典append到列表中。
        return case_lists













